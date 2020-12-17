from os import path, remove
from sys import argv
from pandas import read_excel
from datetime import datetime
from openpyxl import Workbook, load_workbook, styles
from getFiles import file_name
from getLostDate import get_lost_date


"""
对表格进行处理，这里的操作都是简单的excel处理，只是在保存的时候遇到了一些小问题
主要难点在打包上，即如何让数据方便使用这个脚本
使用pyinstaller进行打包，进到当前目录，注意使用虚拟环境，避免打包一些不需要的库，pycharm会自己创建一个虚拟环境
如果需要额外使用自己的虚拟环境的话
pip install pipenv
pipenv install
pipenv shell
进入虚拟环境后，安装脚本需要的库，以及pyinstaller
然后pyinstaller -Fw ./getInformation.py
注意请使用3.7.6及以下版本的python，实测高版本python打包后执行报错
执行完毕后在当前目录下的dist文件夹里会生成一个文件
发给数据后，给文件加上执行权限即可
"""

# 获取目录
# 如果给了参数就使用参数的目录，否则使用默认的目录
# target_dir = ""
# if len(argv) > 1 and path.exists(argv[1]):
#     target_dir = argv[1]
# else:
#     target_dir = "./"

# 直接取当前目录了，通过这种方式去拿当前目录，别的方法可能在打包后拿到错误的目录
target_dir = path.dirname(path.realpath(argv[0])) + "/"

# 删除模板文件
if path.exists(target_dir + "渠道数据整合模板.xlsx"):
    remove(target_dir + "渠道数据整合模板.xlsx")
if path.exists(target_dir + "渠道数据整合模板.csv"):
    remove(target_dir + "渠道数据整合模板.csv")
# 读当前目录下的文件
files = file_name(target_dir)
# 创建模板文件
wb = Workbook()  # 得到一个excel
ws = wb.active  # 得到一个sheet
ws['A1'] = '日期'
ws['B1'] = '渠道'
ws['C1'] = '投放渠道名'
ws['D1'] = '激活率'
ws['E1'] = '操作系统'
ws['F1'] = '供应商'
ws['G1'] = '消耗金额'
ws['H1'] = '下载量'
ws['I1'] = '激活量'
ws['J1'] = '次日留存'
ws['K1'] = 'CPA'
wb.save(target_dir + "渠道数据整合模板.xlsx")

# 目标文件从第二行开始写入，第一行是标题
tar_row = 2
dates = []
for file in files:
    # 这里要设置data_only=True，否则会受到公式的影响
    work_book1 = load_workbook(file, data_only=True)
    # 默认就用第一个sheet
    sheet1 = work_book1.worksheets[0]
    # 源文件也从第二行开始读
    row = 2

    # 文件名格式处理，从
    # /Users/yifan/PycharmProjects/IntegrationChannel/万象新动-马卡龙玩图-oppo-oppo信息流-12.15.xlsx
    # 提取出
    # 万象新动-马卡龙玩图-oppo-oppo信息流-12.15
    simple_file = str("".join(file.rsplit('.', 1)[:-1]).rsplit('/', 1)[1])

    # 设置各列的名字从哪里取的，有的是从文件名里split出来的，有的从表里拿的数据
    supplier = simple_file.split('-')[-5]  # 供应商->供应商
    channel = simple_file.split('-')[-3]  # 渠道->渠道
    channel_name = simple_file.split('-')[-2]  # 渠道包名->投放渠道名
    date = ""  # 日期->日期
    os = ""  # os->操作系统
    download = ""  # 下载->下载量
    activation = ""  # 激活->激活量
    conversion_rate = ""  # 转化率->激活率
    consumption_amount = ""  # 消耗金额->消耗金额
    cpa = ""  # cpa->cpa
    secondary_stay = ""  # 次留->次日留存

    # 开始处理单个表格里的所有行
    while True:
        # 根据日期去判断，只要日期为空，就停止循环
        date = sheet1["A" + str(row)].value
        if date is None:
            break
        # 日期要格式化一下
        if isinstance(date, datetime):
            date = date.strftime('%Y-%m-%d')
            dates.append(date)
        # 拿到数据
        os = sheet1["B" + str(row)].value
        download = sheet1["E" + str(row)].value
        activation = sheet1["F" + str(row)].value
        conversion_rate = sheet1["L" + str(row)].value
        consumption_amount = sheet1["N" + str(row)].value
        cpa = sheet1["O" + str(row)].value
        secondary_stay = sheet1["P" + str(row)].value

        # 拿完一行的数据给row+1，准备拿下一行
        row += 1

        # 拿下一行数据前，要把这一行的数据先写入
        work_book2 = load_workbook(target_dir + "渠道数据整合模板.xlsx")
        ws2 = work_book2.active
        ws2['A' + str(tar_row)] = date
        ws2['B' + str(tar_row)] = channel
        ws2['C' + str(tar_row)] = channel_name
        ws2['D' + str(tar_row)] = conversion_rate
        ws2['E' + str(tar_row)] = os
        ws2['F' + str(tar_row)] = supplier
        ws2['G' + str(tar_row)] = consumption_amount
        ws2['H' + str(tar_row)] = download
        ws2['I' + str(tar_row)] = activation
        ws2['J' + str(tar_row)] = secondary_stay
        ws2['K' + str(tar_row)] = cpa

        # 写入完要给tar_row+1，下次写入的时候换行
        tar_row += 1

        # 每次写完都要保存一下
        work_book2.save(target_dir + "渠道数据整合模板.xlsx")

# 补充剩余日期
lost_dates = get_lost_date(dates)
if lost_dates:
    work_book3 = load_workbook(target_dir + "渠道数据整合模板.xlsx")
    ws3 = work_book3.active
    for lost_date in lost_dates:
        ws3['A' + str(tar_row)] = lost_date
        tar_row += 1
    work_book3.save(target_dir + "渠道数据整合模板.xlsx")


"""根据日期排序
这里其实挺麻烦的，pandas排序首先就是要读excel，读的时候还得设置engine='openpyxl'，否则报错
然后根据列名排序，得设置inplace=True，否则保存了没效果
然后保存一个csv，还得设置index=False，不然第一列会多出来索引，然后得设置编码，否则用excel打开会乱码
然后保存一个excel，也得设置index=False，并且pandas保存后，原来的格式都没了，所以把excel的格式设置都放在了最后
"""
df = read_excel(target_dir + "渠道数据整合模板.xlsx", engine='openpyxl')
df.sort_values(by='日期', ascending=False, inplace=True)
df.to_csv(target_dir + "渠道数据整合模板.csv", index=False, encoding='utf_8_sig')
df.to_excel(target_dir + "渠道数据整合模板.xlsx", index=False)

# 设置excel格式
work_book4 = load_workbook(target_dir + "渠道数据整合模板.xlsx")
ws4 = work_book4.active

# 左对齐，上下居中
alignment = styles.Alignment(horizontal='left', vertical='center')
# 要对每个单元格都设置
for row in ws4.rows:
    for cell in row:
        cell.alignment = alignment
# 设置列宽
ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 20
ws4.column_dimensions['D'].width = 20
ws4.column_dimensions['E'].width = 20
ws4.column_dimensions['F'].width = 20
ws4.column_dimensions['G'].width = 20
ws4.column_dimensions['H'].width = 20
ws4.column_dimensions['I'].width = 20
ws4.column_dimensions['J'].width = 20
ws4.column_dimensions['K'].width = 20

# 重新保存excel
work_book4.save(target_dir + "渠道数据整合模板.xlsx")
